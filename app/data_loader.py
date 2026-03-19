"""
data_loader.py
Parses the Little Alchemist .xlsm file and caches results to JSON.

DATA sheet structure (columns 0-based):
  0  CC_Num     – card number of primary card (populated on every row)
  1  CC_Name    – primary card name (populated on every row)
  2  CC_Rare    – primary card rarity string (populated on every row)
  3  Cmb_Cntr   – total combos for this primary card (populated on every row)
  7  Cmb_Num    – sequence ID of this combination
  8  Cmb_ID     – float lookup key: min(id_A + id_B/1000, id_B + id_A/1000)
  9  CC_A       – first card name
 10  CC_B       – second card name
 11  Cmb_Rare   – rarity-class of this combination (integer 1-4)
 12  Res        – result card name
 13  Res_Rare   – result card rarity (integer 1-4)
 14  BA_0O      – result base attack  when 0 Onyx cards used
 15  BA_1O      – result base attack  when 1 Onyx card  used
 16  BA_2O      – result base attack  when 2 Onyx cards used
 17  BD_0O      – result base defence when 0 Onyx cards used
 18  BD_1O      – result base defence when 1 Onyx card  used
 19  BD_2O      – result base defence when 2 Onyx cards used

USER sheet relevant sections (1-based Excel rows/cols):
  Row  1, cols 20+  : Onyx flags for each deck slot (bool)
  Row  2, cols 20+  : Fused flags ("Yes"/"No") for each deck slot
  Row  3, cols 20+  : Levels for each deck slot (string "3"-"5")
  Row  4, cols 20+  : Numeric card IDs for each deck slot
  Row  5, cols 20+  : Card display names for each deck slot
  Row  8, col 14    : Cached total deck score
  Rows 12-161, col  1  : Library card name
  Rows 12-161, col  2  : Library card level (int)
  Rows 12-161, col  3  : Library card fused ("Yes"/"No")
  Rows 12-161, col  4  : Library card quantity
  Rows 12-161, col 15  : Library card numeric ID
  Rows 12-161, col 18  : Library card Onyx flag (bool)
  Rows  7, col 199  : Current start card name
"""

import json
from pathlib import Path

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Where we cache the parsed combo data (large, read-only game data)
CACHE_FILE = Path(__file__).parent / "combo_data.json"


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _int_safe(v, default=0):
    try:
        return int(v)
    except (TypeError, ValueError):
        return default


def _float_safe(v, default=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _bool_safe(v):
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        return v.strip().lower() in ("yes", "true", "1")
    return bool(v)


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

def _parse_excel(xlsm_path: str) -> dict:
    """Parse the xlsm file and return a plain-dict data structure."""
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required.  Run:  pip install openpyxl"
        )

    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True, data_only=True)

    # ── DATA sheet ──────────────────────────────────────────────────────────
    ws_data = wb["DATA"]
    combo_dict: dict[str, dict] = {}   # str(cmb_id) -> entry
    name_to_id: dict[str, int] = {}    # card_name -> CC_Num
    card_info:  dict[str, dict] = {}   # card_name -> {num, rare, cmb_cntr}

    for row in ws_data.iter_rows(min_row=2, values_only=True):
        # ── Per-card metadata (CC_Num, CC_Name, CC_Rare, Cmb_Cntr are
        #    populated on every row in the DATA sheet) ──────────────────
        cc_num_raw  = row[0]
        cc_name_raw = row[1]
        cc_rare_raw = row[2]
        cmb_cntr_raw = row[3]

        if cc_name_raw and cc_num_raw is not None:
            card_name_clean = str(cc_name_raw).strip()
            cc_num = _int_safe(cc_num_raw, 0)
            if card_name_clean and cc_num:
                name_to_id[card_name_clean] = cc_num
                if card_name_clean not in card_info:
                    card_info[card_name_clean] = {
                        "num":      cc_num,
                        "rare":     str(cc_rare_raw).strip() if cc_rare_raw else "",
                        "cmb_cntr": _int_safe(cmb_cntr_raw, 0),
                    }

        # ── Combo entry ────────────────────────────────────────────────
        cmb_id_raw = row[8]
        card_a = row[9]
        card_b = row[10]
        if cmb_id_raw is None or card_a is None or card_b is None:
            continue

        cmb_id = round(float(cmb_id_raw), 3)

        combo_dict[str(cmb_id)] = {
            "card_a":   card_a,
            "card_b":   card_b,
            "ba":       [row[14] or 0, row[15] or 0, row[16] or 0],
            "bd":       [row[17] or 0, row[18] or 0, row[19] or 0],
            "cmb_rare": _int_safe(row[11], 1),
            "result":   row[12] or "",
            "res_rare": _int_safe(row[13], 1),
        }

    # Sorted unique base-card names for the UI dropdowns (no :Onyx variants)
    base_card_names = sorted(
        name for name in name_to_id.keys() if ":Onyx" not in name
    )

    # ── Advanced Controls sheet ──────────────────────────────────────────
    ws_ac = wb["Advanced Controls"]
    settings = _default_settings()
    for row in ws_ac.iter_rows(min_row=2, max_row=9, values_only=True):
        label = str(row[0]).strip() if row[0] else ""
        # Override column (col C, index 2) takes priority over calculated (B)
        override = row[2]
        calculated = row[1]
        val = override if override is not None else calculated

        if label == "Mode":
            settings["mode"] = _int_safe(val, 1)
        elif "Lowest Combo" in label:
            settings["lcwc"] = _int_safe(val, 35)
        elif "Step Value" in label:
            settings["sv"] = _float_safe(val, 1.0)
        elif "Copy Reduction" in label:
            settings["cr"] = _float_safe(val, 0.8)
        elif "Heroic Mode Attack" in label:
            settings["ab"] = _float_safe(val, 1.5)
        elif "Heroic Mode Defence" in label:
            settings["db"] = _float_safe(val, 0.5)
        elif "Fusion Buff" in label:
            settings["fb"] = _float_safe(val, 2.0)
        elif "Cards to choose" in label:
            settings["n_cards"] = _int_safe(val, 35)

    # Derive combo_name_to_id (sequential card IDs used by the JS scoring engine)
    combo_name_to_id: dict[str, int] = {}
    for key_str, entry in combo_dict.items():
        k = float(key_str)
        id_a = int(k)
        id_b = round((k - id_a) * 1000)
        combo_name_to_id.setdefault(entry["card_a"], id_a)
        combo_name_to_id.setdefault(entry["card_b"], id_b)

    return {
        "combos":           combo_dict,
        "name_to_id":       name_to_id,
        "card_info":        card_info,
        "base_card_names":  base_card_names,
        "combo_name_to_id": combo_name_to_id,
        "settings":         settings,
    }


def _default_settings() -> dict:
    return {
        "mode":    1,    # 1=Sum 2=Attack 3=Defence 4=Heroics
        "lcwc":    35,   # Lowest Combo Worth Counting
        "sv":      1.0,  # Step Value
        "cr":      0.8,  # Copy Reduction
        "fb":      2.0,  # Fusion Buff
        "ab":      1.5,  # Heroic Attack Buff
        "db":      0.5,  # Heroic Defence Buff
        "n_cards": 35,   # Cards to fill automatically
    }


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys
    import glob

    # Accept an explicit path or auto-detect the .xlsm in the parent folder
    if len(sys.argv) > 1:
        xlsm_path = sys.argv[1]
    else:
        candidates = glob.glob(str(Path(__file__).parent.parent / "*.xlsm"))
        if not candidates:
            print("ERROR: No .xlsm file found. Pass path as argument: python data_loader.py <path>")
            sys.exit(1)
        xlsm_path = candidates[0]
        print(f"Auto-detected: {xlsm_path}")

    print(f"Parsing {xlsm_path} ...")
    data = _parse_excel(xlsm_path)

    with open(CACHE_FILE, "w", encoding="utf-8") as fh:
        json.dump(data, fh, indent=2)

    size_kb = CACHE_FILE.stat().st_size // 1024
    print(f"✓ {CACHE_FILE} written  ({size_kb} KB)")
    print(f"  {len(data['combos'])} combos | {len(data['name_to_id'])} cards")
    print("Done. Run build_data.py next to regenerate the JS bundles.")
